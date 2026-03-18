import pandas as pd
import requests
import os
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ─── CONFIGURACIÓN ────────────────────────────────────────────────────────────
QUIP_TOKEN    = "TU_TOKEN_AQUI"                          # https://quip-amazon.com/api/personal-token
QUIP_THREAD_ID = "giFmAddB8FRU"                          # ID del spreadsheet en Quip
CSV_FOLDER    = r"C:\Users\koirod\Desktop\CSV Connect"   # Carpeta donde está el CSV
OUTPUT_FOLDER = r"C:\Users\koirod\Desktop\CSV Connect"   # Carpeta donde se guarda el Excel

HEADERS  = {"Authorization": f"Bearer {QUIP_TOKEN}"}
BASE_URL = "https://platform.quip-amazon.com/1"

# Mapeo de columnas: CSV → Quip
COLUMN_MAP = {
    "OSE":                          "OSE",
    "Job ID":                       "Job ID",
    "Dig Dug":                      "DigDug",
    "Modified OSE Response":        "OSE Response",
    "Correct Event (Connect Data)": "Correct Event",
    "URL Share Point":              "SharePoint",
}

# Normalización de nombres de columnas del Quip
QUIP_COLUMN_RENAME = {
    "OSC":             "OSE",
    "Jod ID":          "Job ID",
    "Digdug":          "DigDug",
    "OSE response":    "OSE Response",
    "Correct event":   "Correct Event",
    "Root cause":      "Root Cause",
    "Sharepoint link": "SharePoint",
}

COLUMNAS_DASHBOARD = ["OSE", "Job ID", "DigDug", "OSE Response", "Correct Event", "SharePoint"]


# ─── FASE 1: LIMPIAR EL CSV ───────────────────────────────────────────────────
def limpiar_csv():
    archivos = [f for f in os.listdir(CSV_FOLDER) if f.endswith(".csv")]
    if not archivos:
        print("❌ No se encontró ningún archivo CSV en la carpeta")
        return None, None

    archivo = sorted(archivos)[-1]
    ruta    = os.path.join(CSV_FOLDER, archivo)
    print(f"📥 Leyendo archivo: {archivo}")

    df = pd.read_csv(ruta)
    print(f"   Filas originales: {len(df)}")

    semana = int(df["Week"].dropna().iloc[0])
    print(f"   Semana detectada: {semana}")

    df = df[df["Real Defects_1"] == 1]
    print(f"   Filas después de filtrar Real Defects_1=1: {len(df)}")

    df = df[list(COLUMN_MAP.keys())].rename(columns=COLUMN_MAP)
    print(f"✅ CSV limpio con {len(df)} filas y {len(df.columns)} columnas")
    return df, semana


# ─── FASE 2: LEER QUIP ────────────────────────────────────────────────────────
def get_quip_spreadsheet(semana):
    url      = f"{BASE_URL}/threads/{QUIP_THREAD_ID}"
    response = requests.get(url, headers=HEADERS)

    if response.status_code != 200:
        print(f"❌ Error conectando al Quip: {response.status_code}")
        return None

    html            = response.json().get("html", "")
    soup            = BeautifulSoup(html, "lxml")
    nombre_hoja     = f"Week {semana} (2026)"

    print(f"   Buscando hoja: {nombre_hoja}")
    tabla = soup.find("table", {"title": nombre_hoja})

    if not tabla:
        print(f"❌ No se encontró la hoja '{nombre_hoja}' en el Quip")
        return None

    all_rows = [[td.get_text(strip=True) for td in tr.find_all("td")]
                for tr in tabla.find_all("tr")]
    all_rows = [r for r in all_rows if r]

    if not all_rows:
        print("❌ No se encontraron filas en la tabla")
        return None

    df = pd.DataFrame(all_rows[1:], columns=all_rows[0])
    df = df.rename(columns=QUIP_COLUMN_RENAME)

    primera_col = df.columns[0]
    if primera_col == "" or all(df[primera_col].astype(str).str.isdigit()):
        df = df.drop(columns=[primera_col])

    print(f"✅ Hoja '{nombre_hoja}' cargada: {len(df)} filas")
    return df


# ─── HELPER: NORMALIZAR SHAREPOINT ───────────────────────────────────────────
def limpiar_sharepoint(valor):
    return str(valor).strip().replace("\u200b", "").strip()


# ─── HELPER: ESCRIBIR HOJA EN EXCEL ──────────────────────────────────────────
def escribir_hoja(ws, df):
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append(["" if pd.isna(v) or str(v) == "nan" else v for v in row])
    for col in ws.columns:
        max_len    = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


# ─── FASE 3: COMPARAR Y GENERAR EXCEL ────────────────────────────────────────
def generar_excel(df_csv, df_quip, semana):
    df_csv["Job ID"]  = df_csv["Job ID"].astype(str).str.strip()
    df_quip["Job ID"] = df_quip["Job ID"].astype(str).str.strip()

    ids_csv   = set(df_csv["Job ID"])
    ids_quip  = set(df_quip["Job ID"])
    comunes   = ids_csv & ids_quip

    # Hoja 1 — Filas nuevas (Job ID no está en el Quip)
    df_nuevas = (df_csv[df_csv["Job ID"].isin(ids_csv - ids_quip)]
                 .copy()
                 .pipe(lambda d: d[[c for c in COLUMNAS_DASHBOARD if c in d.columns]])
                 .sort_values("OSE"))
    print(f"\n➕ Filas nuevas para agregar al Quip: {len(df_nuevas)}")

    # Hoja 2 — SharePoint faltante (está en Quip, Quip no tiene link, CSV sí tiene)
    sp_faltante = []
    for job_id in comunes:
        idx_q = df_quip[df_quip["Job ID"] == job_id].index
        idx_c = df_csv[df_csv["Job ID"] == job_id].index
        if idx_q.empty or idx_c.empty:
            continue
        sp_quip = limpiar_sharepoint(df_quip.loc[idx_q[0], "SharePoint"])
        sp_csv  = limpiar_sharepoint(df_csv.loc[idx_c[0], "SharePoint"])
        if sp_quip in ("", "nan") and sp_csv not in ("", "nan"):
            sp_faltante.append(job_id)

    df_sharepoint = (df_csv[df_csv["Job ID"].isin(set(sp_faltante))]
                     .copy()
                     .pipe(lambda d: d[[c for c in COLUMNAS_DASHBOARD if c in d.columns]])
                     .sort_values("OSE"))
    print(f"🔴 Filas con SharePoint faltante: {len(df_sharepoint)}")

    # Generar Excel
    wb  = Workbook()
    ws1 = wb.active
    ws1.title = f"Week {semana} (2026)"
    if len(df_nuevas) > 0:
        escribir_hoja(ws1, df_nuevas)
    else:
        ws1.append(["No hay filas nuevas para agregar"])

    ws2 = wb.create_sheet(title="SharePoint Faltante")
    if len(df_sharepoint) > 0:
        escribir_hoja(ws2, df_sharepoint)
    else:
        ws2.append(["No hay filas con SharePoint faltante"])

    output_path = os.path.join(OUTPUT_FOLDER, f"Connect_Week_{semana}_2026.xlsx")
    wb.save(output_path)
    print(f"\n✅ Excel guardado en: {output_path}")
    print(f"   Hoja 1 - Week {semana} (2026): {len(df_nuevas)} filas nuevas")
    print(f"   Hoja 2 - SharePoint Faltante: {len(df_sharepoint)} filas")


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    print("\n🚀 Iniciando proceso...\n")

    print("─── FASE 1: Limpiando CSV ───")
    df_csv, semana = limpiar_csv()
    if df_csv is None:
        return

    print("\n─── FASE 2: Cargando Quip ───")
    df_quip = get_quip_spreadsheet(semana)
    if df_quip is None:
        return

    print("\n─── FASE 3: Generando Excel ───")
    generar_excel(df_csv, df_quip, semana)


if __name__ == "__main__":
    main()
